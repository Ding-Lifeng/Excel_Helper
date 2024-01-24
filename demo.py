# -*-coding:utf-8 -*-

"""
# File       : main.py
# Time       ：2024/1/23
# Author     ：丁笠峰
# version    ：python 3.9
"""

# 项目描述：报表修改
# 项目过程：复制文件至目标位置，根据要求修改目标文件

import openpyxl
import pandas as pd

if __name__ == '__main__':
    file_path = "C:\\Users\\25761\\Desktop\\表哥项目\\开票鉴定\\模板文件\\C30.xlsx"

    # # 创建Workbook读取excel的表格
    # Workbook = openpyxl.load_workbook(file_path, data_only=True)
    # sheet_report = Workbook["配合比设计报告"]
    # sheet_certificate = Workbook["合格证"]

    # # 确定配合比设计报告表待修改数据的位置
    # data = [sheet_report.cell(3, 33).value]  # 对应xlsx的AG3-报告日期
    # data.append(sheet_report.cell(4, 33).value)  # 对应xlsx的AG4-报告编号
    # data.append(sheet_report.cell(5, 7).value)  # 对应xlsx的G5-工程名称
    # data.append(sheet_report.cell(6, 7).value)  # 对应xlsx的G6-浇筑部位
    # data.append(sheet_report.cell(8, 7).value)  # 对应xlsx的G8-施工单位
    # data.append(sheet_report.cell(14, 33).value)  # 对应xlsx的AG14-水泥实验编号
    # data.append(sheet_report.cell(15, 33).value)  # 对应xlsx的AG15-砂实验编号
    # data.append(sheet_report.cell(16, 33).value)  # 对应xlsx的AG16-石实验编号
    # data.append(sheet_report.cell(17, 33).value)  # 对应xlsx的AG17-粉煤灰实验编号
    # data.append(sheet_report.cell(18, 33).value)  # 对应xlsx的AG18-矿粉实验编号
    # data.append(sheet_report.cell(19, 33).value)  # 对应xlsx的AG19-泵送剂实验编号
    # print(data)

    # # 确定合格证表待修改数据的位置
    # data = [sheet_certificate.cell(10, 41).value]  # 对应xlsx的AO10-供应数量
    # data.append(sheet_certificate.cell(28, 9).value)  # 对应xlsx的I28-施工地点
    # print(data)

    info_path = "C:\\Users\\25761\\Desktop\\表哥项目\\开票鉴定\\工作表\\info.xlsx"  # 调用信息表
    df = pd.read_excel(info_path)
    i = 0  # 生成表格计数
    for index, row in df.iterrows():
        source_file = openpyxl.load_workbook(file_path, data_only=True)  # 读取源文件
        # 生成转存结果路径
        target_path = "C:\\Users\\25761\\Desktop\\表哥项目\\开票鉴定\\导出文件\\"
        i += 1
        # 转存
        temp_path = target_path + "结果表" + str(i) + ".xlsx"
        source_file.save(temp_path)
        # 创建Workbook读取excel的表格
        Workbook = openpyxl.load_workbook(temp_path, data_only=True)
        sheet_report = Workbook["配合比设计报告"]
        sheet_certificate = Workbook["合格证"]
        # 修改Workbook
        sheet_report.cell(3, 33).value = row['报告日期']
        sheet_report.cell(4, 33).value = row['鉴定编号']
        sheet_report.cell(6, 7).value = row['浇筑部位']
        sheet_report.cell(14, 33).value = row['水泥实验编号']
        sheet_report.cell(15, 33).value = row['砂实验编号']
        sheet_report.cell(16, 33).value = row['石实验编号']
        sheet_report.cell(17, 33).value = row['粉煤灰实验编号']
        sheet_report.cell(18, 33).value = row['矿粉实验编号']
        sheet_report.cell(19, 33).value = row['泵送剂实验编号']
        sheet_certificate.cell(10, 41).value = row['供应数量']
        sheet_certificate.cell(28, 9).value = row['施工地点']
        Workbook.save(temp_path)
